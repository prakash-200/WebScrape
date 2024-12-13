




#
# import requests
# from bs4 import BeautifulSoup
# import openpyxl
# import csv
#
# # Function to apply 1% discount to the price
# def apply_discount(price):
#     try:
#         price = price.replace("Rs.", "").replace(",", "").strip()  # Remove "Rs." and commas
#         price_float = float(price)  # Convert the cleaned price string to float
#         discounted_price = price_float - (price_float * 0.01)  # 1% discount
#         return round(discounted_price, 2)
#     except ValueError as e:
#         print(f"Error applying discount to price: {price} | Error: {e}")
#         return None  # Return None if conversion or calculation fails
#
# # Function to convert specifications to structured HTML format
# def convert_spec_to_html(specification_div):
#     html_output = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     for li_tag in specification_div.find_all('li'):
#         divs = li_tag.find_all('div')
#         if len(divs) == 2:
#             key = divs[0].text.strip()
#             value = divs[1].text.strip()
#             html_output += f'<tr><td style="padding: 8px; text-align: left;">{key}</td><td style="padding: 8px; text-align: left;">{value}</td></tr>'
#     html_output += '</table>'
#     return html_output
#
# # Function to update image URLs to use 300x300
# def update_image_url(url):
#     if url and "600x600" in url:
#         return url.replace("600x600", "300x300")
#     return url
#
# # Function to extract product details from a URL
# def extract_product_details(url, category):
#     try:
#         response = requests.get(url)
#         response.raise_for_status()
#
#         soup = BeautifulSoup(response.text, 'html.parser')
#         title_tag = soup.find('h1')
#         title = title_tag.text.strip() if title_tag else "No Title"
#
#         product_image = None
#         image_div = soup.find('div', class_='theme-product-detail-image-inner')
#         if image_div:
#             picture_tag = image_div.find('picture')
#             if picture_tag:
#                 img_tag = picture_tag.find('img')
#                 if img_tag and 'src' in img_tag.attrs:
#                     product_image = update_image_url(f"https://techgurustore.in/{img_tag['src']}")
#
#         additional_images = []
#         thumbnails_div = soup.find('div', class_='theme-product-detail-thumbnail')
#         if thumbnails_div:
#             for picture_tag in thumbnails_div.find_all('picture'):
#                 img_tag = picture_tag.find('img')
#                 if img_tag and 'src' in img_tag.attrs:
#                     additional_images.append(
#                         update_image_url(f"https://techgurustore.in/{img_tag['src']}")
#                     )
#
#         price_span = soup.find('span', class_='theme-product-price')
#         price = price_span.text.strip() if price_span else None
#         discounted_price = apply_discount(price) if price else None
#
#         regular_price_span = soup.find('span', class_='theme-product-old-price')
#         regular_price = regular_price_span.text.strip() if regular_price_span else None
#
#         short_description = ""
#         short_desc_div = soup.find('div', class_='theme-product-short-description')
#         if short_desc_div:
#             short_description = short_desc_div.get_text(strip=True)
#
#         details_span = soup.find('div', class_='theme-product-info-content')
#         product_details = details_span.text.strip() if details_span else None
#
#         specification_html = None
#         specification_div = soup.find('div', class_='theme-prod-specification-table')
#         if specification_div:
#             specification_html = convert_spec_to_html(specification_div)
#
#         return {
#             'Category': category,
#             'URL': url,
#             'Product Image': product_image,
#             'Additional Images': ', '.join(additional_images),
#             'Title': title,
#             'Price': price,
#             'Discounted Price': discounted_price,
#             'Regular Price': regular_price,
#             'Short Description': short_description,
#             'Product Details': product_details,
#             'Specification (HTML)': specification_html
#         }
#
#     except requests.exceptions.RequestException as e:
#         print(f"Error fetching URL {url}: {e}")
#         return None
#
# # Function to read categories and URLs from Excel
# def read_categories_and_urls_from_excel(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2) if row[0].value and row[1].value]
#     return data
#
# # Function to save data to CSV
# def save_to_csv(data, output_file):
#     with open(output_file, mode='w', newline='', encoding='utf-8') as file:
#         writer = csv.DictWriter(file, fieldnames=data[0].keys())
#         writer.writeheader()
#         writer.writerows(data)
#
# # Main script
# def main():
#     input_excel = 'ProductLinks.xlsx'  # Replace with your input Excel file
#     output_csv = 'ProductData.csv'  # Specify the output CSV file
#
#     category_url_data = read_categories_and_urls_from_excel(input_excel)
#     all_product_data = []
#
#     for category, url in category_url_data:
#         print(f"Extracting data for category '{category}' from: {url}")
#         product_data = extract_product_details(url, category)
#         if product_data:
#             all_product_data.append(product_data)
#
#     if all_product_data:
#         save_to_csv(all_product_data, output_csv)
#         print(f"Data saved to {output_csv}")
#     else:
#         print("No data extracted.")
#
# # Run the script
# if __name__ == "__main__":
#     main()





















#
#
#
# import requests
# from bs4 import BeautifulSoup
# import openpyxl
# import csv
#
# # Function to apply 1% discount to the price
# def apply_discount(price):
#     try:
#         # If the price is a range, return it directly
#         if ' - ' in price:
#             return price  # Return the whole range as the discount price
#
#         # Otherwise, process the price normally
#         price = price.replace("Rs.", "").replace(",", "").strip()  # Remove "Rs." and commas
#         price_float = float(price)  # Convert the cleaned price string to float
#         discounted_price = price_float - (price_float * 0.01)  # 1% discount
#         return round(discounted_price, 2)
#     except ValueError as e:
#         print(f"Error applying discount to price: {price} | Error: {e}")
#         return price  # If conversion fails, return the original price
#
# # Function to convert specifications to structured HTML format
# def convert_spec_to_html(specification_div):
#     html_output = '<table border="1" style="border-collapse: collapse; width: 100%;">'
#     for li_tag in specification_div.find_all('li'):
#         divs = li_tag.find_all('div')
#         if len(divs) == 2:
#             key = divs[0].text.strip()
#             value = divs[1].text.strip()
#             html_output += f'<tr><td style="padding: 8px; text-align: left;">{key}</td><td style="padding: 8px; text-align: left;">{value}</td></tr>'
#     html_output += '</table>'
#     return html_output
#
# # Function to update image URLs to use 300x300
# def update_image_url(url):
#     if url and "600x600" in url:
#         return url.replace("600x600", "300x300")
#     return url
#
# # Function to extract product details from a URL
# def extract_product_details(url, category):
#     try:
#         response = requests.get(url)
#         response.raise_for_status()
#
#         soup = BeautifulSoup(response.text, 'html.parser')
#         title_tag = soup.find('h1')
#         title = title_tag.text.strip() if title_tag else "No Title"
#
#         product_image = None
#         image_div = soup.find('div', class_='theme-product-detail-image-inner')
#         if image_div:
#             picture_tag = image_div.find('picture')
#             if picture_tag:
#                 img_tag = picture_tag.find('img')
#                 if img_tag and 'src' in img_tag.attrs:
#                     product_image = update_image_url(f"https://techgurustore.in/{img_tag['src']}")
#
#         additional_images = []
#         thumbnails_div = soup.find('div', class_='theme-product-detail-thumbnail')
#         if thumbnails_div:
#             for picture_tag in thumbnails_div.find_all('picture'):
#                 img_tag = picture_tag.find('img')
#                 if img_tag and 'src' in img_tag.attrs:
#                     additional_images.append(
#                         update_image_url(f"https://techgurustore.in/{img_tag['src']}")
#                     )
#
#         price_span = soup.find('span', class_='theme-product-price')
#         price = price_span.text.strip() if price_span else None
#         discounted_price = apply_discount(price) if price else price  # If discount can't be applied, use original price
#
#         regular_price_span = soup.find('span', class_='theme-product-old-price')
#         regular_price = regular_price_span.text.strip() if regular_price_span else None
#
#         short_description = ""
#         short_desc_div = soup.find('div', class_='theme-product-short-description')
#         if short_desc_div:
#             short_description = short_desc_div.get_text(strip=True)
#
#         details_span = soup.find('div', class_='theme-product-info-content')
#         product_details = details_span.text.strip() if details_span else None
#
#         specification_html = None
#         specification_div = soup.find('div', class_='theme-prod-specification-table')
#         if specification_div:
#             specification_html = convert_spec_to_html(specification_div)
#
#         return {
#             'Category': category,
#             'URL': url,
#             'Product Image': product_image,
#             'Additional Images': ', '.join(additional_images),
#             'Title': title,
#             'Price': price,
#             'Discounted Price': discounted_price,
#             'Regular Price': regular_price,
#             'Short Description': short_description,
#             'Product Details': product_details,
#             'Specification (HTML)': specification_html
#         }
#
#     except requests.exceptions.RequestException as e:
#         print(f"Error fetching URL {url}: {e}")
#         return None
#
# # Function to read categories and URLs from Excel
# def read_categories_and_urls_from_excel(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2) if row[0].value and row[1].value]
#     return data
#
# # Function to save data to CSV
# def save_to_csv(data, output_file):
#     with open(output_file, mode='w', newline='', encoding='utf-8') as file:
#         writer = csv.DictWriter(file, fieldnames=data[0].keys())
#         writer.writeheader()
#         writer.writerows(data)
#
# # Main script
# def main():
#     input_excel = 'ProductLinks.xlsx'  # Replace with your input Excel file
#     output_csv = 'ProductData.csv'  # Specify the output CSV file
#
#     category_url_data = read_categories_and_urls_from_excel(input_excel)
#     all_product_data = []
#
#     for category, url in category_url_data:
#         print(f"Extracting data for category '{category}' from: {url}")
#         product_data = extract_product_details(url, category)
#         if product_data:
#             all_product_data.append(product_data)
#
#     if all_product_data:
#         save_to_csv(all_product_data, output_csv)
#         print(f"Data saved to {output_csv}")
#     else:
#         print("No data extracted.")
#
# # Run the script
# if __name__ == "__main__":
#     main()



import requests
from bs4 import BeautifulSoup
import openpyxl
import csv
import os

# Function to apply 1% discount to the price
def apply_discount(price):
    try:
        # If the price is a range, return it directly
        if ' - ' in price:
            return price  # Return the whole range as the discount price

        # Otherwise, process the price normally
        price = price.replace("Rs.", "").replace(",", "").strip()  # Remove "Rs." and commas
        price_float = float(price)  # Convert the cleaned price string to float
        discounted_price = price_float - (price_float * 0.01)  # 1% discount
        return round(discounted_price, 2)
    except ValueError as e:
        print(f"Error applying discount to price: {price} | Error: {e}")
        return price  # If conversion fails, return the original price

# Function to convert specifications to structured HTML format
def convert_spec_to_html(specification_div):
    html_output = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    for li_tag in specification_div.find_all('li'):
        divs = li_tag.find_all('div')
        if len(divs) == 2:
            key = divs[0].text.strip()
            value = divs[1].text.strip()
            html_output += f'<tr><td style="padding: 8px; text-align: left;">{key}</td><td style="padding: 8px; text-align: left;">{value}</td></tr>'
    html_output += '</table>'
    return html_output

# Function to update image URLs to use 300x300
def update_image_url(url):
    if url and "600x600" in url:
        return url.replace("600x600", "300x300")
    return url

# Function to extract product details from a URL
def extract_product_details(url, category):
    try:
        response = requests.get(url)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        title_tag = soup.find('h1')
        title = title_tag.text.strip() if title_tag else "No Title"

        product_image = None
        image_div = soup.find('div', class_='theme-product-detail-image-inner')
        if image_div:
            picture_tag = image_div.find('picture')
            if picture_tag:
                img_tag = picture_tag.find('img')
                if img_tag and 'src' in img_tag.attrs:
                    product_image = update_image_url(f"https://techgurustore.in/{img_tag['src']}")

        additional_images = []
        thumbnails_div = soup.find('div', class_='theme-product-detail-thumbnail')
        if thumbnails_div:
            for picture_tag in thumbnails_div.find_all('picture'):
                img_tag = picture_tag.find('img')
                if img_tag and 'src' in img_tag.attrs:
                    additional_images.append(
                        update_image_url(f"https://techgurustore.in/{img_tag['src']}")
                    )

        price_span = soup.find('span', class_='theme-product-price')
        price = price_span.text.strip() if price_span else None
        discounted_price = apply_discount(price) if price else price  # If discount can't be applied, use original price

        regular_price_span = soup.find('span', class_='theme-product-old-price')
        regular_price = regular_price_span.text.strip() if regular_price_span else None

        short_description = ""
        short_desc_div = soup.find('div', class_='theme-product-short-description')
        if short_desc_div:
            short_description = short_desc_div.get_text(strip=True)

        details_span = soup.find('div', class_='theme-product-info-content')
        product_details = details_span.text.strip() if details_span else None

        specification_html = None
        specification_div = soup.find('div', class_='theme-prod-specification-table')
        if specification_div:
            specification_html = convert_spec_to_html(specification_div)

        return {
            'Category': category,
            'URL': url,
            'Product Image': product_image,
            'Additional Images': ', '.join(additional_images),
            'Title': title,
            'Price': price,
            'Discounted Price': discounted_price,
            'Regular Price': regular_price,
            'Short Description': short_description,
            'Product Details': product_details,
            'Specification (HTML)': specification_html
        }

    except requests.exceptions.RequestException as e:
        print(f"Error fetching URL {url}: {e}")
        return None

# Function to read categories and URLs from Excel
def read_categories_and_urls_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2) if row[0].value and row[1].value]
    return data

# Function to save data to CSV, ensuring header is only written once
def save_to_csv(data, output_file):
    # Check if file exists to decide whether to write header
    file_exists = os.path.exists(output_file)
    with open(output_file, mode='a', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=data[0].keys())
        # Write header only if file does not exist
        if not file_exists:
            writer.writeheader()
        writer.writerows(data)

# Main script
def main():
    input_excel = 'ProductLinks.xlsx'  # Replace with your input Excel file
    output_csv = 'ProductData.csv'  # Specify the output CSV file

    category_url_data = read_categories_and_urls_from_excel(input_excel)
    all_product_data = []

    for category, url in category_url_data:
        print(f"Extracting data for category '{category}' from: {url}")
        product_data = extract_product_details(url, category)
        if product_data:
            all_product_data.append(product_data)
            save_to_csv([product_data], output_csv)  # Save product data immediately after extraction

    if all_product_data:
        print(f"All data saved to {output_csv}")
    else:
        print("No data extracted.")

# Run the script
if __name__ == "__main__":
    main()
